from celery import shared_task
from openpyxl import load_workbook
from datetime import datetime
import os
from django.apps import apps

@shared_task
def ingest_customer_data():
    from django.conf import settings
    file_path = os.path.join(settings.BASE_DIR, 'data', 'customer_data.xlsx')
    
    if not os.path.exists(file_path):
        return "customer_data.xlsx not found"
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            
            customer_id = int(row[0])
            first_name = str(row[1]) if row[1] else ""
            last_name = str(row[2]) if row[2] else ""
            age = int(row[3]) if row[3] else 30
            phone_number = int(row[4]) if row[4] else 0
            monthly_salary = int(row[5]) if row[5] else 0
            approved_limit = int(row[6]) if row[6] else 0
            current_debt = 0.0
            
            Customer = apps.get_model('customers', 'Customer')
            Customer.objects.update_or_create(
                customer_id=customer_id,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'age': age,
                    'phone_number': phone_number,
                    'monthly_salary': monthly_salary,
                    'approved_limit': approved_limit,
                    'current_debt': current_debt,
                }
            )
        
        return "Customer data ingested successfully"
    except Exception as e:
        return f"Error ingesting customer data: {str(e)}"


@shared_task
def ingest_loan_data():
    from django.conf import settings
    file_path = os.path.join(settings.BASE_DIR, 'data', 'loan_data.xlsx')
    
    if not os.path.exists(file_path):
        return "loan_data.xlsx not found"
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        
        customer_id = int(row[0])
        
        Customer = apps.get_model('customers', 'Customer')
        if not Customer.objects.filter(customer_id=customer_id).exists():
            continue
        
        loan_id = int(row[1])
        loan_amount = float(row[2])
        tenure = int(row[3])
        interest_rate = float(row[4])
        monthly_repayment = float(row[5])
        emis_paid_on_time = int(row[6])
        start_date_str = row[7]
        end_date_str = row[8]
        
        if isinstance(start_date_str, str):
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = start_date_str
        
        if isinstance(end_date_str, str):
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = end_date_str
        
        customer = Customer.objects.get(customer_id=customer_id)
        
        Loan = apps.get_model('loans', 'Loan')
        Loan.objects.update_or_create(
            loan_id=loan_id,
            defaults={
                'customer': customer,
                'loan_amount': loan_amount,
                'tenure': tenure,
                'interest_rate': interest_rate,
                'monthly_repayment': monthly_repayment,
                'emis_paid_on_time': emis_paid_on_time,
                'start_date': start_date,
                'end_date': end_date,
            }
        )
    
    return "Loan data ingested successfully"
